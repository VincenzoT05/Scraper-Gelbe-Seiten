from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import time
import csv
import os
from openpyxl import Workbook, load_workbook

# Configurazione Chrome Driver con webdriver-manager
options = Options()
# options.add_argument('--headless')  # scommenta per esecuzione senza GUI
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)
wait = WebDriverWait(driver, 30)  # timeout più lungo

start_url = "https://www.gelbeseiten.de/suche/cocktailbars/bundesweit"
driver.get(start_url)

csv_filename = 'cocktailbars.csv'
excel_filename = 'cocktailbars.xlsx'

def save_to_csv(row, fieldnames):
    file_exists = os.path.isfile(csv_filename)
    with open(csv_filename, mode='a', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        if not file_exists:
            writer.writeheader()
        writer.writerow(row)

def save_to_excel(row):
    if os.path.exists(excel_filename):
        workbook = load_workbook(excel_filename)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Nome', 'Indirizzo', 'Telefono', 'Sito Web', 'Email', 'Link Dettaglio'])
    sheet.append([row['Nome'], row['Indirizzo'], row['Telefono'], row['Sito Web'], row['Email'], row['Link Dettaglio']])
    workbook.save(excel_filename)

def load_all_results():
    print("Caricamento di tutti i risultati cliccando il link 'Mehr Anzeigen'...")
    previous_count = 0
    while True:
        bars = driver.find_elements(By.CSS_SELECTOR, 'article.mod-Treffer')
        current_count = len(bars)
        print(f"Articoli trovati finora: {current_count}")
        if current_count == previous_count:
            print("Nessun nuovo articolo caricato; termino caricamento.")
            break
        previous_count = current_count
        try:
            load_more_link = wait.until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, 'a#mod-LoadMore--button.mod-LoadMore--button')
                )
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", load_more_link)
            driver.execute_script("window.scrollBy(0, -100);")  # scrolla sopra per evitare overlay
            time.sleep(1)
            driver.execute_script("arguments[0].click();", load_more_link)
            time.sleep(3)
        except Exception as e:
            print(f"Errore cliccando il link 'Mehr Anzeigen' o link non più presente: {e}")
            break

fieldnames = ['Nome', 'Indirizzo', 'Telefono', 'Sito Web', 'Email', 'Link Dettaglio']

print("Caricamento iniziale pagina e risultati completi...")
time.sleep(3)

load_all_results()

print("Parsing pagina con tutti i risultati caricati...")
soup = BeautifulSoup(driver.page_source, 'html.parser')
bars = soup.find_all('article', class_='mod-Treffer')

MAX_RETRIES = 3

print(f"Trovati {len(bars)} bar. Inizio estrazione dati...")

for index, bar in enumerate(bars, start=1):
    try:
        nome_tag = bar.find('h2', class_='mod-Treffer__name')
        nome = nome_tag.get_text(strip=True) if nome_tag else ''

        indirizzo_tag = bar.find('div', class_='mod-AdresseKompakt__adress-text')
        indirizzo = indirizzo_tag.get_text(separator=' ', strip=True) if indirizzo_tag else ''

        telefono_tag = bar.find('a', class_='mod-TelefonnummerKompakt__phoneNumber')
        telefono = telefono_tag.get_text(strip=True) if telefono_tag else ''

        dettaglio_link_tag = bar.find('a', href=True)
        dettaglio_link = dettaglio_link_tag['href'] if dettaglio_link_tag else ''
        if dettaglio_link.startswith('/'):
            dettaglio_link = "https://www.gelbeseiten.de" + dettaglio_link

        sito_web = ''
        email = ''

        if dettaglio_link:
            # Retry per apertura pagina dettaglio
            for attempt in range(1, MAX_RETRIES + 1):
                try:
                    driver.get(dettaglio_link)
                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.aktionsleiste-button')))
                    break
                except Exception as e:
                    print(f"Tentativo {attempt} per caricare {dettaglio_link} fallito: {e}")
                    if attempt == MAX_RETRIES:
                        print(f"Salto la pagina dettaglio di '{nome}' dopo {MAX_RETRIES} tentativi.")
                        sito_web = ''
                        email = ''
                        break
                    time.sleep(3)

            if sito_web == '' and attempt < MAX_RETRIES:
                detail_soup = BeautifulSoup(driver.page_source, 'html.parser')

                sito_tag = detail_soup.select_one('div.aktionsleiste-button a[href]')
                if sito_tag:
                    sito_web = sito_tag['href']

                email_button = detail_soup.find('div', id='email_versenden')
                if email_button:
                    data_link = email_button.get('data-link', '')
                    if data_link.startswith('mailto:'):
                        email = data_link.split(':')[1].split('?')[0]

            # Retry per tornare indietro
            for back_attempt in range(1, MAX_RETRIES + 1):
                try:
                    driver.back()
                    wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'mod-Treffer')))
                    time.sleep(1)
                    break
                except Exception as e:
                    print(f"Errore al tornare indietro, tentativo {back_attempt}: {e}")
                    if back_attempt == MAX_RETRIES:
                        print("Impossibile tornare indietro, chiudo browser e termino.")
                        driver.quit()
                        exit(1)
                    time.sleep(2)

        row = {
            'Nome': nome,
            'Indirizzo': indirizzo,
            'Telefono': telefono,
            'Sito Web': sito_web,
            'Email': email,
            'Link Dettaglio': dettaglio_link
        }

        print(f"[{index}/{len(bars)}] Salvo: {row}")
        save_to_csv(row, fieldnames)
        save_to_excel(row)

    except Exception as e:
        print(f"Errore durante l'elaborazione del bar '{nome}': {e}")

driver.quit()
print("Scraping completato! File CSV e Excel salvati.")
